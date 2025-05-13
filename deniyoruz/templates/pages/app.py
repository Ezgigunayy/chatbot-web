from flask import Flask, render_template, request, redirect, session
import MySQLdb
import pandas as pd
from flask_mysqldb import MySQL
import os
import mysql.connector
from config import Config
import re
from flask import send_file
from openpyxl import Workbook
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "your_secret_key"

app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'chatbot'

mysql = MySQL(app)

@app.route('/', methods=['GET', 'POST'])
def login():
    msg = ''
    if request.method == 'POST' and 'Student_number' in request.form and 'Password' in request.form:
        student_number = request.form['Student_number']
        password = request.form['Password']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM users WHERE Student_number = %s AND Password = %s', (student_number, password))
        account = cursor.fetchone()
        if account:
            session['loggedin'] = True
            session['Student_number'] = account['Student_number']
            msg = 'Başarıyla giriş yapıldı!'
            return redirect('/chatbot')
        else:
            msg = 'Geçersiz kullanıcı adı veya şifre!'
    return render_template('login.html', msg=msg)


@app.route('/register', methods=['GET', 'POST'])
def register():
    msg = ''
    if request.method == 'POST' and 'Student_number' in request.form and 'Password' in request.form and 'name' in request.form:
        student_number = request.form['Student_number']
        password = request.form['Password']
        name = request.form['name']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM users WHERE Student_number = %s', (student_number,))
        account = cursor.fetchone()
        if account:
            msg = 'Bu öğrenci numarası zaten kayıtlı!'
        elif not re.match(r'[0-9]+', student_number):
            msg = 'Öğrenci numarası sadece rakamlardan oluşmalıdır!'
        elif not student_number or not password or not name:
            msg = 'Tüm alanlar doldurulmalıdır!'
        else:
            cursor.execute('INSERT INTO users (Student_number, Password, Name) VALUES (%s, %s, %s)', (student_number, password, name))
            mysql.connection.commit()
            msg = 'Başarıyla kayıt oldunuz!'
            return redirect('/')
    return render_template('register.html', msg=msg)

@app.route('/export_data_to_excel', methods=['POST'])
def export_data_to_excel():
    file_path = "C:/Users/Acer/Documents/GitHub/As-lYedek/deniyoruz/questions.xlsx"

    # Formdan gelen veriler
    question = request.form.get('question')
    answer = request.form.get('answer')
    sentiment = request.form.get('Feedback')
    topic = request.form.get('topic')
    student_number = session.get('Student_number')

    if not (question and answer and sentiment and topic and student_number):
        return "Eksik veri gönderildi!", 400

    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Student_number", "question", "answer", "Feedback", "topic"])

    ws = wb.active
    ws.append([student_number, question, answer, sentiment, topic])
    wb.save(file_path)

    return "Veriler başarıyla Excel'e kaydedildi!"


@app.route('/chatbot', methods=['GET', 'POST'])
def chatbot():
    if 'loggedin' in session:
        if request.method == 'POST':
            question = request.form.get('question')
            answer = request.form.get('answer')
            sentiment = request.form.get('Feedback')
            topic = request.form.get('topic')
            student_number = session.get('Student_number')
            
            cursor = mysql.connection.cursor()
            try:
                cursor.execute(
                    'INSERT INTO questions (ogrenci_ID, questions, answers, Feedback, topic) VALUES (%s, %s, %s, %s, %s)',
                    (student_number, question, answer, sentiment, topic)
                )
                mysql.connection.commit()
                msg = 'Kayıt başarıyla eklendi!'
                export_data_to_excel()
                
            except Exception as e:
                msg = f'Kayıt hatası: {e}'
            return render_template('chatbot.html', msg=msg)
        return render_template('chatbot.html')
    return redirect('/')



@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    msg = ''
    if request.method == 'POST':
        student_number = request.form.get('Student_number')
        new_password = request.form.get('new_password')
        
        # Öğrenci numarası ve yeni şifre kontrolü
        if not student_number.strip() or not new_password.strip():
            msg = 'Öğrenci numarası ve şifre boş olamaz!'
            return render_template('reset_password.html', msg=msg)

        # Öğrenci numarasına ait kayıt olup olmadığını kontrol et
        cursor = mysql.connection.cursor()
        cursor.execute('SELECT * FROM users WHERE Student_number = %s', (student_number,))
        existing_record = cursor.fetchone()

        if existing_record:
            # Yeni şifreyi güncelle
            cursor.execute('UPDATE users SET Password = %s WHERE Student_number = %s', (new_password, student_number))
            mysql.connection.commit()
            msg = 'Şifre başarıyla sıfırlandı!'
            return redirect('/')  # Başarıyla sıfırlama sonrası login sayfasına yönlendir
        else:
            msg = 'Bu öğrenci numarası ile kayıt bulunamadı!'
    
    return render_template('reset_password.html', msg=msg)

@app.route('/logout')
def logout():
    session.pop('loggedin', None)
    session.pop('Student_number', None)
    return redirect('/')


if __name__ == "__main__":
    app.run(debug=True)
